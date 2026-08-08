"""Generate a well-formatted Excel validation spreadsheet for peer raters.

Creates an .xlsx with:
- Instructions sheet with rubrics and examples
- Ratings sheet with 130 papers, Likert dropdowns, frozen headers
- Rubric Quick Reference sheet (printable)
"""

from __future__ import annotations

import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import (
    Alignment, Border, Font, PatternFill, Side,
)
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation

from .paths import RunPaths, build_run_paths, ensure_run_dirs


def _sample_papers(df: pd.DataFrame, n_target: int = 130, seed: int = 42) -> pd.DataFrame:
    """Stratified sample of papers for validation."""
    # Exclude pilot papers
    pilot_path = Path(__file__).resolve().parents[1] / "runs" / "latin_science_2026" / "indicators" / "pilot_likert_sample.csv"
    if pilot_path.exists():
        pilot_ids = set(pd.read_csv(pilot_path)["corpus_id"])
        df = df[~df["id"].isin(pilot_ids)].copy()

    df["year_cohort"] = pd.cut(
        df["publication_year"], bins=[2019, 2022, 2023, 2024, 2026],
        labels=["2020-22", "2023", "2024", "2025-26"],
    )

    rng = np.random.default_rng(seed)
    alloc = {"2020-22": 10, "2023": 25, "2024": 40, "2025-26": 55}
    parts = []

    for cohort, n_cohort in alloc.items():
        pool = df[df["year_cohort"] == cohort]
        if len(pool) <= n_cohort:
            parts.append(pool.copy())
            continue

        pool["T_tercile"] = pd.qcut(pool["axis_t_technology"], 3, labels=["Low", "Mid", "High"], duplicates="drop")
        pool["G_tercile"] = pd.qcut(pool["axis_g_governance"], 3, labels=["Low", "Mid", "High"], duplicates="drop")

        cells = list(pool.groupby(["T_tercile", "G_tercile"], observed=False))
        per_cell = max(1, n_cohort // len(cells))
        picked_list = []
        for _, cell in cells:
            n = min(len(cell), per_cell)
            if n > 0:
                picked_list.append(cell.sample(n, random_state=seed))

        picked = pd.concat(picked_list, ignore_index=True)
        if len(picked) < n_cohort:
            remaining = pool[~pool["id"].isin(picked["id"])]
            extra_n = min(n_cohort - len(picked), len(remaining))
            if extra_n > 0:
                picked = pd.concat([picked, remaining.sample(extra_n, random_state=seed)], ignore_index=True)
        parts.append(picked)

    sample = pd.concat(parts, ignore_index=True).drop_duplicates(subset=["id"])
    if len(sample) > n_target:
        sample = sample.sample(n_target, random_state=seed)

    # Generate stable form IDs
    rng_ids = np.random.default_rng(123)
    letters = list("ABCDEFGHJKLMNPQRSTUVWXYZ")
    form_ids = ["V" + "".join(rng_ids.choice(letters, 5).tolist()) for _ in range(len(sample))]
    sample["form_id"] = form_ids
    return sample


def generate(paths: RunPaths, n_papers: int = 130) -> Path:
    """Generate the validation Excel file.

    Returns the path to the generated .xlsx file.
    """
    try:
        sys.stdout.reconfigure(encoding="utf-8")
    except (AttributeError, ValueError):
        pass

    ensure_run_dirs(paths)

    # Load data
    corpus = pd.read_csv(paths.corpus_paper_path)
    axis = pd.read_csv(paths.indicators_dir / "axis_scores.csv")
    df = corpus.merge(axis[["id", "axis_t_technology", "axis_g_governance"]], on="id", how="left")
    df = df[df["decision"] == "include"].copy()

    sample = _sample_papers(df, n_target=n_papers)
    print(f"Sampled {len(sample)} papers for validation")

    # ---- Build workbook ---------------------------------------------------
    wb = Workbook()

    # Reusable styles
    header_fill = PatternFill(start_color="1B3A5C", end_color="1B3A5C", fill_type="solid")
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    section_font = Font(name="Calibri", size=12, bold=True, color="1B3A5C")
    body_font = Font(name="Calibri", size=11)
    thin_border = Border(
        left=Side(style="thin", color="D0D0D0"),
        right=Side(style="thin", color="D0D0D0"),
        top=Side(style="thin", color="D0D0D0"),
        bottom=Side(style="thin", color="D0D0D0"),
    )
    even_fill = PatternFill(start_color="F2F7FC", end_color="F2F7FC", fill_type="solid")
    odd_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")
    score_fill = PatternFill(start_color="FFF8E1", end_color="FFF8E1", fill_type="solid")

    # ===== Sheet 1: Instructions =====
    ws_info = wb.active
    ws_info.title = "Instructions"
    ws_info.column_dimensions["A"].width = 5
    ws_info.column_dimensions["B"].width = 100

    def write_section(ws, row, title, lines, fill=None, font=None):
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
        cell = ws.cell(row=row, column=1, value=title)
        cell.font = font or section_font
        if fill:
            cell.fill = fill
        row += 1
        for line in lines:
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=2)
            cell = ws.cell(row=row, column=1, value=line)
            cell.font = body_font
            cell.alignment = Alignment(wrap_text=True)
            row += 1
        row += 1
        return row

    row = 1
    row = write_section(ws_info, row, "LATIN.SCIENCE 2026 - VALIDATION FORM INSTRUCTIONS", [], fill=header_fill, font=header_font)
    row = write_section(ws_info, row, "What you are rating", [
        "You will rate 130 paper titles and abstracts on TWO dimensions using 1-5 Likert scales.",
        "Rate ONLY based on the title and abstract provided. Do not search for the full text.",
        "If the information is genuinely insufficient to decide, type 'X' in the Insufficient Info column.",
        "This is a blind validation: you cannot see authors, journal, year, or automated scores.",
    ])
    row = write_section(ws_info, row, "AXIS T - Technological Specificity", [
        "How specifically does the paper engage with NAMED generative AI models?",
        "",
        "  1 = Generic AI/ML: NO named generative model/system is central.",
        '      Ex: "Artificial intelligence can help with peer review."',
        "  2 = GenAI/LLM appears GENERICALLY: no named model as central object.",
        '      Ex: "Large language models offer new possibilities for scientific writing."',
        "  3 = MIXED case or INCIDENTAL mention of a product/model.",
        '      Ex: "We used tools such as ChatGPT for..." (but paper is not about ChatGPT).',
        "  4 = A named model/family is IMPORTANT to the study.",
        '      Ex: A study comparing reviewer performance with and without GPT-4 assistance.',
        "  5 = Study CENTRALLY evaluates, compares, or discusses named model(s).",
        '      Ex: "Evaluating GPT-4, Claude, and Gemini for abstract screening accuracy."',
        "",
        "Named models include: ChatGPT, GPT-3.5, GPT-4, GPT-4o, Claude, Gemini, LLaMA, Copilot, Bard, etc.",
    ])
    row = write_section(ws_info, row, "AXIS G - Governance/Workflow Orientation", [
        "Is the paper primarily about SUPPORTING scholarly tasks, or about GOVERNING/REGULATING AI?",
        "",
        "  1 = CENTRALLY about supporting research/writing/review/synthesis TASKS.",
        '      Ex: "An AI tool to automate systematic review screening."',
        "  2 = PREDOMINANTLY workflow support, with secondary mention of rules/concerns.",
        '      Ex: "LLMs for manuscript drafting, with a brief note on ethical considerations."',
        "  3 = BALANCED or indeterminate focus between tools and governance.",
        '      Ex: "AI in peer review: opportunities and ethical challenges."',
        "  4 = PREDOMINANTLY about integrity, authorship, disclosure, policy, or safeguards.",
        '      Ex: "Institutional policies for detecting AI-generated student submissions."',
        "  5 = CENTRALLY about governance/integrity/policy; task support is secondary.",
        '      Ex: "A framework for academic integrity in the age of generative AI."',
        "",
        "Governance topics: academic integrity, plagiarism detection, authorship ethics,",
        "AI-text detection, disclosure policies, institutional guardrails, publication ethics.",
        "Workflow topics: writing assistance, screening automation, literature search,",
        "peer review support, reference management, data extraction, evidence synthesis tools.",
    ])
    row = write_section(ws_info, row, "HOW TO FILL THE RATINGS SHEET", [
        "1. Go to the Ratings tab at the bottom.",
        "2. For each paper, read the title and abstract.",
        "3. Select a score (1-5) from the dropdown in columns E (T) and G (G).",
        '4. If you cannot decide, type "X" in column F (T) or H (G) for Insufficient Info.',
        "5. Use column I (Notes) for any comments - optional but helpful for edge cases.",
        "6. Save and return the file when done.",
        "",
        "Tips:",
        "- Use the FULL 1-5 range. Do NOT default to 3.",
        "- Trust your first impression after reading the abstract.",
        '- "Insufficient info" should be RARE - most abstracts are adequate.',
        "- There are no right or wrong answers; we are measuring agreement, not accuracy.",
    ])

    # ===== Sheet 2: Ratings =====
    ws = wb.create_sheet("Ratings")

    col_widths = {"A": 5, "B": 10, "C": 55, "D": 75, "E": 10, "F": 13, "G": 10, "H": 13, "I": 30}
    for col_letter, width in col_widths.items():
        ws.column_dimensions[col_letter].width = width

    headers = ["#", "Form ID", "Title", "Abstract", "T (1-5)", "T Insuf.\nInfo?", "G (1-5)", "G Insuf.\nInfo?", "Notes"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
    ws.row_dimensions[1].height = 35

    # Data validation dropdowns
    dv_T = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    dv_T.error = "Please select a value from 1 to 5"
    dv_T.errorTitle = "Invalid score"
    dv_T.prompt = "Select T score (1-5)"
    dv_T.promptTitle = "Technological Specificity"
    ws.add_data_validation(dv_T)

    dv_G = DataValidation(type="list", formula1='"1,2,3,4,5"', allow_blank=True)
    dv_G.error = "Please select a value from 1 to 5"
    dv_G.errorTitle = "Invalid score"
    dv_G.prompt = "Select G score (1-5)"
    dv_G.promptTitle = "Governance Orientation"
    ws.add_data_validation(dv_G)

    # Data rows
    for i, (_, paper) in enumerate(sample.iterrows()):
        r = i + 2
        row_fill = even_fill if i % 2 == 0 else odd_fill

        # Sequence number
        cell = ws.cell(row=r, column=1, value=i + 1)
        cell.font = Font(name="Calibri", size=9, color="888888")
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = row_fill

        # Form ID
        cell = ws.cell(row=r, column=2, value=paper["form_id"])
        cell.font = Font(name="Consolas", size=9, bold=True)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.fill = row_fill

        # Title
        cell = ws.cell(row=r, column=3, value=str(paper["title"]))
        cell.font = Font(name="Calibri", size=10)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.fill = row_fill

        # Abstract (truncate if very long)
        abstract = str(paper["abstract"])
        if len(abstract) > 2000:
            abstract = abstract[:1997] + "..."
        cell = ws.cell(row=r, column=4, value=abstract)
        cell.font = Font(name="Calibri", size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.fill = row_fill

        # T score (dropdown)
        cell = ws.cell(row=r, column=5)
        cell.fill = score_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        dv_T.add(cell)

        # T insufficient
        cell = ws.cell(row=r, column=6, value="")
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

        # G score (dropdown)
        cell = ws.cell(row=r, column=7)
        cell.fill = score_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border
        dv_G.add(cell)

        # G insufficient
        cell = ws.cell(row=r, column=8, value="")
        cell.fill = row_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

        # Notes
        cell = ws.cell(row=r, column=9, value="")
        cell.font = Font(name="Calibri", size=9)
        cell.alignment = Alignment(wrap_text=True, vertical="top")
        cell.fill = row_fill
        cell.border = thin_border

        ws.row_dimensions[r].height = 80

    # Freeze panes and auto-filter
    ws.freeze_panes = "C2"
    ws.auto_filter.ref = f"A1:I{len(sample) + 1}"

    # Rubric reminder at bottom
    legend_row = len(sample) + 3
    ws.merge_cells(start_row=legend_row, start_column=1, end_row=legend_row, end_column=9)
    legend_text = (
        "T: 1=Generic AI/ML  2=GenAI generic  3=Mixed/incidental  4=Named model important  5=Named model central"
        "    |    "
        "G: 1=Workflow central  2=Mostly workflow  3=Balanced  4=Mostly governance  5=Governance central"
    )
    cell = ws.cell(row=legend_row, column=1, value=legend_text)
    cell.font = Font(name="Calibri", size=9, italic=True, color="666666")
    cell.alignment = Alignment(horizontal="center")

    # ===== Sheet 3: Rubric Quick Reference =====
    ws_ref = wb.create_sheet("Rubric Reference")
    ws_ref.column_dimensions["A"].width = 5
    ws_ref.column_dimensions["B"].width = 18
    ws_ref.column_dimensions["C"].width = 60
    ws_ref.column_dimensions["D"].width = 5
    ws_ref.column_dimensions["E"].width = 18
    ws_ref.column_dimensions["F"].width = 60

    ws_ref.merge_cells("A1:C1")
    cell = ws_ref.cell(row=1, column=1, value="AXIS T - Technological Specificity")
    cell.font = Font(name="Calibri", size=14, bold=True, color="1B3A5C")

    t_rubric = [
        ("1", "Generic AI/ML", "No named generative model or system is central. Paper discusses AI/ML broadly."),
        ("2", "GenAI generic", "GenAI/LLMs appear generically. No named model as central object."),
        ("3", "Mixed / incidental", "Mixed case or incidental mention of a specific product or model."),
        ("4", "Named model important", "A named model/family is important to the study (ChatGPT, GPT-4, Claude...)."),
        ("5", "Named model central", "Study centrally evaluates, compares, or discusses one or more named models."),
    ]
    for j, (score, label, desc) in enumerate(t_rubric):
        r = 3 + j
        ws_ref.cell(row=r, column=1, value=score).font = Font(name="Calibri", size=12, bold=True, color="1B3A5C")
        ws_ref.cell(row=r, column=2, value=label).font = Font(name="Calibri", size=11, bold=True)
        ws_ref.cell(row=r, column=3, value=desc).font = Font(name="Calibri", size=11)
        ws_ref.cell(row=r, column=3).alignment = Alignment(wrap_text=True)
        ws_ref.row_dimensions[r].height = 22
        if j % 2 == 0:
            for c in range(1, 4):
                ws_ref.cell(row=r, column=c).fill = even_fill

    ws_ref.merge_cells("E1:F1")
    cell = ws_ref.cell(row=1, column=5, value="AXIS G - Governance/Workflow Orientation")
    cell.font = Font(name="Calibri", size=14, bold=True, color="1B3A5C")

    g_rubric = [
        ("1", "Workflow central", "Centrally about supporting research, writing, review, or synthesis tasks."),
        ("2", "Mostly workflow", "Predominantly workflow support, with secondary mention of rules/concerns."),
        ("3", "Balanced", "Balanced or indeterminate focus between tools/support and governance."),
        ("4", "Mostly governance", "Predominantly about integrity, authorship, disclosure, policy, or safeguards."),
        ("5", "Governance central", "Centrally about governance, integrity, or policy. Task support is secondary."),
    ]
    for j, (score, label, desc) in enumerate(g_rubric):
        r = 3 + j
        ws_ref.cell(row=r, column=5, value=score).font = Font(name="Calibri", size=12, bold=True, color="1B3A5C")
        ws_ref.cell(row=r, column=6, value=label).font = Font(name="Calibri", size=11, bold=True)
        ws_ref.cell(row=r, column=7, value=desc).font = Font(name="Calibri", size=11)
        ws_ref.cell(row=r, column=7).alignment = Alignment(wrap_text=True)
        ws_ref.row_dimensions[r].height = 22
        if j % 2 == 0:
            for c in range(5, 8):
                ws_ref.cell(row=r, column=c).fill = even_fill

    # ---- Save ---------------------------------------------------------------
    out_path = paths.indicators_dir / "validation_ratings_130.xlsx"
    wb.save(out_path)
    print(f"Saved: {out_path}")
    print(f"  Sheets: {wb.sheetnames}")
    print(f"  Papers: {len(sample)}")
    print(f"  Features: Likert dropdowns, frozen headers, alternating rows, rubric reference")
    return out_path


if __name__ == "__main__":
    paths = build_run_paths("runs/latin_science_2026")
    generate(paths)
